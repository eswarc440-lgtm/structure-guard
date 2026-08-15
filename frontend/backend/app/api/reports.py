"""
Report generation endpoints for SIMRAS.
Supports PDF, CSV, and Excel export formats.
"""

import io
import csv
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors

from app.database.database import get_db
from app.models.infrastructure import InfrastructureAsset
from app.models.inspections import Inspection
from app.models.maintenance import Maintenance
from app.models.gis_and_analytics import RiskAssessment

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])


def format_currency(value: Optional[float]) -> str:
    """Format currency value with rupee symbol."""
    if value is None:
        return "N/A"
    return f"₹{value:,.2f}"


def format_date(value: Optional[datetime]) -> str:
    """Format date value."""
    if value is None:
        return "N/A"
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    return str(value)


@router.get("/asset/{asset_id}/pdf", response_class=Response)
async def export_asset_pdf(asset_id: str, db: Session = Depends(get_db)):
    """
    Generate PDF report for a specific asset.
    Includes asset details, inspection history, maintenance records, and risk assessment.
    """
    # Fetch asset data
    asset = db.query(InfrastructureAsset).filter(
        InfrastructureAsset.id == asset_id
    ).first()
    
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    # Fetch related data
    inspections = db.query(Inspection).filter(
        Inspection.asset_id == asset_id
    ).order_by(Inspection.inspection_date.desc()).all()
    
    maintenance_records = db.query(Maintenance).filter(
        Maintenance.asset_id == asset_id
    ).order_by(Maintenance.planned_start_date.desc()).all()
    
    risk_assessment = db.query(RiskAssessment).filter(
        RiskAssessment.asset_id == asset_id
    ).order_by(RiskAssessment.created_at.desc()).first()
    
    # Create PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        alignment=1,  # Center
    )
    story.append(Paragraph("SIMRAS Asset Report", title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Asset Information
    story.append(Paragraph("Asset Information", styles['Heading2']))
    asset_data = [
        ["Field", "Value"],
        ["Asset ID", asset.id],
        ["Asset Type", asset.asset_type or "N/A"],
        ["Location", asset.location or "N/A"],
        ["Condition Score", f"{asset.health_score:.1f}/100" if asset.health_score else "N/A"],
        ["Age", f"{asset.design_life or 0} years" if asset.design_life else "N/A"],
        ["Latitude", f"{asset.latitude:.6f}" if asset.latitude else "N/A"],
        ["Longitude", f"{asset.longitude:.6f}" if asset.longitude else "N/A"],
    ]
    
    asset_table = Table(asset_data, colWidths=[2*inch, 3*inch])
    asset_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    story.append(asset_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Risk Assessment
    if risk_assessment:
        story.append(Paragraph("Risk Assessment", styles['Heading2']))
        risk_data = [
            ["Metric", "Value"],
            ["Risk Score", f"{risk_assessment.risk_score:.1f}/100"],
            ["Risk Level", risk_assessment.risk_level.upper()],
            ["Confidence", f"{risk_assessment.confidence_score:.1%}"],
            ["Valid Until", format_date(risk_assessment.valid_until)],
        ]
        
        risk_table = Table(risk_data, colWidths=[2*inch, 3*inch])
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ffe6e6')]),
        ]))
        story.append(risk_table)
        
        if risk_assessment.risk_explanation:
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph("<b>Risk Explanation:</b>", styles['Normal']))
            story.append(Paragraph(risk_assessment.risk_explanation, styles['Normal']))
        
        story.append(Spacer(1, 0.3*inch))
    
    # Inspections
    if inspections:
        story.append(PageBreak())
        story.append(Paragraph(f"Recent Inspections ({len(inspections)})", styles['Heading2']))
        
        for idx, inspection in enumerate(inspections[:5], 1):
            insp_data = [
                ["Date", format_date(inspection.inspection_date)],
                ["Type", inspection.inspection_type or "General"],
                ["Condition Score", f"{inspection.condition_score:.1f}/100" if inspection.condition_score else "N/A"],
                ["Crack Score", f"{inspection.crack_score:.1f}/100" if inspection.crack_score else "N/A"],
                ["Corrosion Score", f"{inspection.corrosion_score:.1f}/100" if inspection.corrosion_score else "N/A"],
                ["Remarks", inspection.remarks or "No remarks"],
            ]
            
            story.append(Paragraph(f"Inspection #{idx}", styles['Heading3']))
            
            insp_table = Table(insp_data, colWidths=[2*inch, 3*inch])
            insp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecfdf5')]),
            ]))
            story.append(insp_table)
            story.append(Spacer(1, 0.2*inch))
    
    # Maintenance
    if maintenance_records:
        story.append(PageBreak())
        story.append(Paragraph(f"Maintenance Records ({len(maintenance_records)})", styles['Heading2']))
        
        for idx, maint in enumerate(maintenance_records[:5], 1):
            maint_data = [
                ["Type", maint.maintenance_type or "Unspecified"],
                ["Status", maint.status.upper()],
                ["Priority", maint.priority or "Normal"],
                ["Planned Start", format_date(maint.planned_start_date)],
                ["Actual Cost", format_currency(maint.actual_cost)],
                ["Contractor", maint.assigned_contractor or "Unassigned"],
            ]
            
            story.append(Paragraph(f"Maintenance #{idx}", styles['Heading3']))
            
            maint_table = Table(maint_data, colWidths=[2*inch, 3*inch])
            maint_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3e8ff')]),
            ]))
            story.append(maint_table)
            story.append(Spacer(1, 0.2*inch))
    
    # Build PDF
    doc.build(story)
    pdf_buffer.seek(0)
    
    return Response(
        content=pdf_buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=asset_{asset_id}_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@router.get("/assets/csv", response_class=Response)
async def export_assets_csv(db: Session = Depends(get_db)):
    """
    Export all assets to CSV format.
    Includes basic asset info, condition scores, and location data.
    """
    assets = db.query(InfrastructureAsset).limit(10000).all()
    
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    
    # Header
    writer.writerow([
        "Asset ID",
        "Asset Type",
        "Location",
        "District",
        "Health Score",
        "Condition Level",
        "Latitude",
        "Longitude",
        "Design Life (years)",
        "Flood Risk Value",
        "Annual Rainfall (mm)",
        "Created Date",
    ])
    
    # Data rows
    for asset in assets:
        writer.writerow([
            asset.id,
            asset.asset_type or "",
            asset.location or "",
            asset.district or "",
            f"{asset.health_score:.2f}" if asset.health_score else "",
            "Good" if asset.health_score and asset.health_score >= 80 else "Fair" if asset.health_score and asset.health_score >= 60 else "Poor",
            f"{asset.latitude:.6f}" if asset.latitude else "",
            f"{asset.longitude:.6f}" if asset.longitude else "",
            asset.design_life or "",
            f"{asset.flood_risk_value:.2f}" if asset.flood_risk_value else "",
            f"{asset.rainfall_mm:.2f}" if asset.rainfall_mm else "",
            format_date(asset.created_at) if hasattr(asset, 'created_at') and asset.created_at else "",
        ])
    
    return Response(
        content=csv_buffer.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=assets_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


@router.get("/summary/xlsx", response_class=Response)
async def export_summary_xlsx(db: Session = Depends(get_db)):
    """
    Export comprehensive summary to Excel with multiple sheets.
    Includes assets, inspections, maintenance, and risk summary.
    """
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Assets Summary Sheet
    ws_assets = wb.create_sheet("Assets")
    assets = db.query(InfrastructureAsset).limit(1000).all()
    
    headers = ["Asset ID", "Type", "City", "State", "Health Score", "Condition", "Latitude", "Longitude"]
    ws_assets.append(headers)
    
    for cell in ws_assets[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for asset in assets:
        condition = "Good" if asset.health_score and asset.health_score >= 80 else "Fair" if asset.health_score and asset.health_score >= 60 else "Poor"
        ws_assets.append([
            asset.id,
            asset.asset_type or "",
            asset.location or "",
            asset.district or "",
            asset.health_score or "",
            condition,
            asset.latitude or "",
            asset.longitude or "",
        ])
        
        for cell in ws_assets[ws_assets.max_row]:
            cell.border = border
    
    ws_assets.column_dimensions['A'].width = 15
    ws_assets.column_dimensions['B'].width = 12
    ws_assets.column_dimensions['C'].width = 12
    ws_assets.column_dimensions['D'].width = 12
    ws_assets.column_dimensions['E'].width = 12
    ws_assets.column_dimensions['F'].width = 12
    ws_assets.column_dimensions['G'].width = 12
    ws_assets.column_dimensions['H'].width = 12
    
    # Inspections Summary Sheet
    ws_insp = wb.create_sheet("Inspections")
    inspections = db.query(Inspection).order_by(Inspection.inspection_date.desc()).limit(500).all()
    
    insp_headers = ["Asset ID", "Inspection Date", "Type", "Condition Score", "Crack Score", "Corrosion Score", "Status"]
    ws_insp.append(insp_headers)
    
    for cell in ws_insp[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for insp in inspections:
        ws_insp.append([
            insp.asset_id,
            format_date(insp.inspection_date),
            insp.inspection_type or "",
            f"{insp.condition_score:.2f}" if insp.condition_score else "",
            f"{insp.crack_score:.2f}" if insp.crack_score else "",
            f"{insp.corrosion_score:.2f}" if insp.corrosion_score else "",
            "Complete",
        ])
        
        for cell in ws_insp[ws_insp.max_row]:
            cell.border = border
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_insp.column_dimensions[col].width = 14
    
    # Maintenance Summary Sheet
    ws_maint = wb.create_sheet("Maintenance")
    maintenance = db.query(Maintenance).order_by(Maintenance.planned_start_date.desc()).limit(500).all()
    
    maint_headers = ["Asset ID", "Type", "Status", "Priority", "Planned Start", "Actual Cost", "Contractor"]
    ws_maint.append(maint_headers)
    
    for cell in ws_maint[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for maint in maintenance:
        ws_maint.append([
            maint.asset_id,
            maint.maintenance_type or "",
            maint.status or "",
            maint.priority or "",
            format_date(maint.planned_start_date),
            format_currency(maint.actual_cost),
            maint.assigned_contractor or "",
        ])
        
        for cell in ws_maint[ws_maint.max_row]:
            cell.border = border
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_maint.column_dimensions[col].width = 14
    
    # Risk Summary Sheet
    ws_risk = wb.create_sheet("Risk Assessment")
    risk_records = db.query(RiskAssessment).order_by(RiskAssessment.risk_score.desc()).limit(500).all()
    
    risk_headers = ["Asset ID", "Risk Score", "Risk Level", "Condition Factor", "Age Factor", "Maintenance Factor", "Confidence"]
    ws_risk.append(risk_headers)
    
    for cell in ws_risk[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for risk in risk_records:
        ws_risk.append([
            risk.asset_id,
            f"{risk.risk_score:.2f}" if risk.risk_score else "",
            risk.risk_level.upper() if risk.risk_level else "",
            f"{risk.condition_factor:.2f}" if risk.condition_factor else "",
            f"{risk.age_factor:.2f}" if risk.age_factor else "",
            f"{risk.maintenance_factor:.2f}" if risk.maintenance_factor else "",
            f"{risk.confidence_score:.2%}" if risk.confidence_score else "",
        ])
        
        for cell in ws_risk[ws_risk.max_row]:
            cell.border = border
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_risk.column_dimensions[col].width = 14
    
    # Save to buffer
    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    
    return Response(
        content=xlsx_buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=simras_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )
